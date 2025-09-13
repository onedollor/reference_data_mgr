"""
Excel Form Processing Utility
Validates and extracts configuration data from completed Excel forms
"""

import os
import re
import logging
from typing import Dict, Any, Tuple, List, Optional
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
class ExcelProcessor:
    """Processes and validates completed Excel configuration forms"""

    def __init__(self, logger: Optional[Any] = None):
        """Initialize the Excel processor"""
        self.logger = logger

        # Valid options for form validation
        self.valid_delimiters = [',', ';', '|', 'Tab']
        self.valid_encodings = ['utf-8', 'utf-16', 'iso-8859-1', 'cp1252']
        self.valid_processing_modes = ['fullload', 'append']
        self.valid_yes_no = ['Yes', 'No']

    def validate_form(self, excel_path: str) -> Tuple[bool, Dict[str, Any], List[str]]:
        """
        Validate completed Excel form and extract configuration

        Args:
            excel_path: Path to the completed Excel form

        Returns:
            Tuple of (is_valid, config_data, error_messages)
        """
        try:
            # Check if Excel file exists and is accessible
            if not os.path.exists(excel_path):
                return False, {}, [f"Excel file not found: {excel_path}"]

            # Extract configuration from Excel
            config_data = self.extract_configuration(excel_path)

            # Validate the extracted configuration
            is_valid, errors = self._validate_configuration(config_data)

            if is_valid:
                print(f"Excel form validation successful: {excel_path}")
            else:
                print(f"Excel form validation failed: {excel_path}, errors: {errors}")

            return is_valid, config_data, errors

        except Exception as e:
            error_msg = f"Failed to validate Excel form {excel_path}: {str(e)}"
            print(error_msg)
            return False, {}, [error_msg]

    def extract_configuration(self, excel_path: str) -> Dict[str, Any]:
        """
        Extract configuration data from Excel form

        Args:
            excel_path: Path to the Excel form file

        Returns:
            Dictionary containing extracted configuration
        """
        try:
            # Load the Excel workbook
            wb = load_workbook(excel_path, data_only=True)
            ws = wb.active

            # Extract configuration from Excel cells
            config = self._parse_excel_content(ws)

            # Set default values for missing fields
            default_config = {
                'delimiter': ',',
                'text_qualifier': '"',
                'encoding': 'utf-8',
                'has_headers': True,
                'processing_mode': 'fullload',
                'is_reference_data': True,
                'table_name': '',
                'confirmed': False,
                'processed_by': '',
                'process_date': ''
            }

            # Merge with extracted config
            default_config.update(config)

            return default_config

        except Exception as e:
            print(f"Failed to extract configuration from Excel {excel_path}: {str(e)}")
            raise

    def _parse_excel_content(self, ws) -> Dict[str, Any]:
        """
        Parse Excel worksheet to extract form data from specific cells
        """
        config = {}

        try:
            # Define the expected structure based on our form generation
            # We'll scan for labels and get values from adjacent cells

            cell_mappings = {
                # Format section mappings
                'Delimiter:': 'delimiter',
                'Encoding:': 'encoding',
                'Text Qualifier:': 'text_qualifier',
                'Has Headers:': 'has_headers',

                # Processing section mappings
                'Mode:': 'processing_mode',
                'Create Config Record:': 'is_reference_data',
                'Table Name:': 'table_name',

                # Confirmation section mappings
                'I Confirm Processing:': 'confirmed',
                'Processed By:': 'processed_by',
                'Date/Time:': 'process_date'
            }

            # Scan through cells to find labels and extract adjacent values
            for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=3):
                cell_a, cell_b, cell_c = row

                if cell_a.value and str(cell_a.value).strip() in cell_mappings:
                    field_name = cell_mappings[str(cell_a.value).strip()]
                    field_value = cell_b.value

                    if field_value is not None:
                        # Process the field value based on field type
                        config[field_name] = self._process_field_value(field_name, field_value)

            # Convert processing mode and reference data to internal format
            if 'processing_mode' in config:
                if config['processing_mode'] == 'fullload':
                    config['processing_mode'] = 'fullload'
                elif config['processing_mode'] == 'append':
                    config['processing_mode'] = 'append'
                elif config['processing_mode'] == 'Select Mode':
                    # User hasn't selected a mode yet
                    config['processing_mode'] = ''

            if 'is_reference_data' in config:
                config['is_reference_data'] = (config['is_reference_data'] == 'Yes')

            if 'has_headers' in config:
                config['has_headers'] = (config['has_headers'] == 'Yes')

            if 'confirmed' in config:
                config['confirmed'] = (config['confirmed'] == 'Yes')

            # Convert delimiter display back to actual character
            if 'delimiter' in config:
                delimiter_map = {'Tab': '\t', ',': ',', ';': ';', '|': '|'}
                config['delimiter'] = delimiter_map.get(config['delimiter'], config['delimiter'])

            # Handle text qualifier
            if 'text_qualifier' in config:
                if config['text_qualifier'] == 'None':
                    config['text_qualifier'] = ''

            # Ensure processed_by and process_date are strings
            if 'processed_by' in config and config['processed_by']:
                config['processed_by'] = str(config['processed_by']).strip()

            if 'process_date' in config and config['process_date']:
                config['process_date'] = str(config['process_date']).strip()

        except Exception as e:
            print(f"Failed to parse Excel content: {str(e)}")
            # Return empty config, validation will catch missing required fields

        return config

    def _process_field_value(self, field_name: str, field_value: Any) -> Any:
        """Process field value based on field type"""
        if field_value is None:
            return ''

        # Convert to string and strip whitespace
        value = str(field_value).strip()

        # Handle specific field types
        if field_name in ['confirmed', 'has_headers', 'is_reference_data']:
            return value  # Will be converted to boolean later

        return value

    def _validate_configuration(self, config: Dict[str, Any]) -> Tuple[bool, List[str]]:
        """
        Validate extracted configuration data

        Args:
            config: Configuration dictionary to validate

        Returns:
            Tuple of (is_valid, error_messages)
        """
        errors = []

        # Required field validation
        required_fields = ['delimiter', 'encoding', 'processing_mode', 'confirmed']
        for field in required_fields:
            if field not in config or config[field] is None or config[field] == '':
                errors.append(f"Required field missing or empty: {field}")

        # Confirmation validation - most critical
        if not config.get('confirmed', False):
            errors.append("Final confirmation must be set to 'Yes' to proceed with processing")

        # Delimiter validation
        delimiter = config.get('delimiter', '')
        if delimiter and delimiter not in [',', ';', '|', '\t']:
            errors.append(f"Invalid delimiter '{delimiter}'. Valid options: comma (,), semicolon (;), pipe (|), tab")

        # Encoding validation
        encoding = config.get('encoding', '')
        if encoding and encoding.lower() not in [e.lower() for e in self.valid_encodings]:
            errors.append(f"Invalid encoding '{encoding}'. Valid options: {', '.join(self.valid_encodings)}")

        # Processing mode validation
        processing_mode = config.get('processing_mode', '')
        if not processing_mode or processing_mode not in ['fullload', 'append']:
            errors.append("Processing mode is required. Please select 'fullload' or 'append'")

        # Table name validation (if provided)
        table_name = config.get('table_name', '')
        if table_name and not re.match(r'^[A-Za-z][A-Za-z0-9_]*$', table_name):
            errors.append(f"Invalid table name '{table_name}'. Must start with letter and contain only letters, numbers, and underscores")

        # Reference data consistency check
        if config.get('is_reference_data', False) and not table_name:
            errors.append("Table name is required when 'Create Config Record' is set to Yes")

        # Processed by validation
        if config.get('confirmed', False) and not config.get('processed_by', '').strip():
            errors.append("'Processed By' field is required when processing is confirmed")

        # Processing mode and reference data consistency
        if config.get('processing_mode') == 'fullload' and config.get('is_reference_data', False):
            # This is valid but add informational note
            print("fullload mode with reference data table will truncate existing data")

        is_valid = len(errors) == 0
        return is_valid, errors

    def get_processing_configuration(self, excel_path: str) -> Dict[str, Any]:
        """
        Get processing configuration for use with ReferenceDataAPI.process_file

        Args:
            excel_path: Path to validated Excel form

        Returns:
            Dictionary formatted for ReferenceDataAPI processing
        """
        is_valid, config, errors = self.validate_form(excel_path)

        if not is_valid:
            raise ValueError(f"Excel form validation failed: {', '.join(errors)}")

        # Extract CSV file path from Excel path
        excel_file = Path(excel_path)
        csv_path = excel_file.parent / f"{excel_file.stem.replace('_config', '')}.csv"

        # Format for ReferenceDataAPI
        processing_config = {
            'csv_file_path': str(csv_path),
            'load_type': config['processing_mode'],
            'is_reference_data': config['is_reference_data'],
            'table_name': config.get('table_name', ''),
            'format_overrides': {
                'delimiter': config['delimiter'],
                'encoding': config['encoding'],
                'has_headers': config['has_headers'],
                'text_qualifier': config.get('text_qualifier', '"')
            },
            'processed_by': config.get('processed_by', 'Unknown'),
            'process_date': config.get('process_date', datetime.now().isoformat())
        }

        return processing_config

    def is_excel_ready_for_processing(self, excel_path: str) -> bool:
        """
        Quick check to see if Excel is ready for processing (confirmed)

        Args:
            excel_path: Path to Excel form file

        Returns:
            True if Excel is confirmed and ready for processing
        """
        try:
            config = self.extract_configuration(excel_path)
            return config.get('confirmed', False)
        except Exception as e:
            print(f"Error checking Excel readiness: {str(e)}")
            return False