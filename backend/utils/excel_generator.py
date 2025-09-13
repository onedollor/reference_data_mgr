"""
Excel Form Generation Utility
Creates interactive Excel configuration forms with dropdowns, validation, and formatting
"""

import os
import logging
from typing import Dict, Any, Optional
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from .file_handler import FileHandler
class ExcelFormGenerator:
    """Generates interactive Excel configuration forms from CSV analysis"""

    def __init__(self, logger: Optional[Any] = None):
        """Initialize the Excel form generator"""
        self.logger = logger
        self.file_handler = FileHandler()

    def generate_form(self, csv_path: str, format_data: Dict[str, Any], workflow_id: Optional[str] = None) -> str:
        """
        Generate interactive Excel configuration form from CSV analysis

        Args:
            csv_path: Path to the CSV file being processed
            format_data: Dictionary containing CSV format detection results
            workflow_id: Optional workflow identifier to include in the form

        Returns:
            Path to generated Excel form file
        """
        try:
            # Generate Excel file path
            excel_path = self._generate_excel_path(csv_path)

            # Create the Excel form with interactive fields
            self._create_interactive_form(excel_path, csv_path, format_data, workflow_id)

            print(f"Generated interactive Excel form: {excel_path}")
            return excel_path

        except Exception as e:
            print(f"Failed to generate Excel form for {csv_path}: {str(e)}")
            raise

    def _generate_excel_path(self, csv_path: str) -> str:
        """Generate Excel file path based on CSV file path"""
        csv_file = Path(csv_path)
        excel_name = f"{csv_file.stem}_config.xlsx"

        # Place Excel in same directory as CSV file
        excel_path = csv_file.parent / excel_name
        return str(excel_path)

    def _create_interactive_form(self, excel_path: str, csv_path: str, format_data: Dict[str, Any], workflow_id: Optional[str] = None):
        """Create the interactive Excel form with data validation and formatting"""

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "CSV Processing Config"

        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15

        # Define styles
        title_font = Font(size=16, bold=True, color="2F5597")
        section_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        section_font = Font(color="FFFFFF", bold=True, size=12)
        label_font = Font(bold=True, size=11)
        warning_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        warning_font = Font(color="FFFFFF", bold=True, size=11)

        border_thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Current row tracker
        row = 1

        # Title
        ws.merge_cells(f'A{row}:C{row}')
        title_cell = ws[f'A{row}']
        title_cell.value = "Reference Data Processing Configuration"
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center')
        row += 2

        # File Information Section
        row = self._add_file_info_section(ws, csv_path, format_data, row, section_fill, section_font, label_font, workflow_id)
        row += 1

        # Format Detection Section
        row = self._add_format_section(ws, format_data, row, section_fill, section_font, label_font, border_thin)
        row += 1

        # Processing Configuration Section
        row = self._add_processing_section(ws, csv_path, row, section_fill, section_font, label_font, border_thin)
        row += 1

        # Confirmation Section
        row = self._add_confirmation_section(ws, row, warning_fill, warning_font, label_font, border_thin)
        row += 1

        # Instructions Section
        self._add_instructions_section(ws, row)

        # Save the workbook
        wb.save(excel_path)

    def _add_file_info_section(self, ws, csv_path: str, format_data: Dict[str, Any], row: int,
                              section_fill, section_font, label_font, workflow_id: Optional[str] = None) -> int:
        """Add file information section"""

        # Section header
        ws.merge_cells(f'A{row}:C{row}')
        header_cell = ws[f'A{row}']
        header_cell.value = "File Information"
        header_cell.fill = section_fill
        header_cell.font = section_font
        header_cell.alignment = Alignment(horizontal='center')
        row += 1

        # File details
        file_path = Path(csv_path)
        info_items = [
            ("File Name:", file_path.name),
            ("File Path:", str(file_path)),
            ("File Size (MB):", f"{format_data.get('file_size_mb', 'Unknown')}"),
            ("Detected Rows:", str(format_data.get('sample_rows', 'Unknown'))),
            ("Detection Time:", datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        ]
        
        # Add workflow_id if provided
        if workflow_id:
            info_items.insert(0, ("Workflow ID:", workflow_id))

        for label, value in info_items:
            ws[f'A{row}'].value = label
            ws[f'A{row}'].font = label_font
            ws[f'B{row}'].value = value
            row += 1

        return row

    def _add_format_section(self, ws, format_data: Dict[str, Any], row: int,
                           section_fill, section_font, label_font, border_thin) -> int:
        """Add detected format section with editable fields"""

        # Section header
        ws.merge_cells(f'A{row}:C{row}')
        header_cell = ws[f'A{row}']
        header_cell.value = "Detected CSV Format (Editable)"
        header_cell.fill = section_fill
        header_cell.font = section_font
        header_cell.alignment = Alignment(horizontal='center')
        row += 1

        # Delimiter field with dropdown
        ws[f'A{row}'].value = "Delimiter:"
        ws[f'A{row}'].font = label_font
        delimiter_cell = ws[f'B{row}']

        # Map detected delimiter to display value
        detected_delimiter = format_data.get('column_delimiter', ',')
        delimiter_display = {',': ',', ';': ';', '|': '|', '\t': 'Tab'}.get(detected_delimiter, ',')
        delimiter_cell.value = delimiter_display
        delimiter_cell.border = border_thin

        # Add delimiter validation with enhanced settings
        delimiter_validation = DataValidation(type="list", formula1='",;|Tab"', showDropDown=False)
        delimiter_validation.error = "Please select a valid delimiter: comma (,), semicolon (;), pipe (|), or Tab"
        delimiter_validation.errorTitle = "Invalid Delimiter"
        delimiter_validation.prompt = "Click dropdown arrow to select delimiter"
        delimiter_validation.promptTitle = "Delimiter Selection"
        delimiter_validation.showErrorMessage = True
        delimiter_validation.showInputMessage = True
        ws.add_data_validation(delimiter_validation)
        delimiter_validation.add(delimiter_cell)
        row += 1

        # Encoding field with dropdown
        ws[f'A{row}'].value = "Encoding:"
        ws[f'A{row}'].font = label_font
        encoding_cell = ws[f'B{row}']

        # Map detected encoding to valid options
        detected_encoding = format_data.get('encoding', 'utf-8')
        encoding_map = {
            'ascii': 'utf-8',
            'utf-8': 'utf-8',
            'latin-1': 'iso-8859-1',
            'cp1252': 'cp1252',
            'utf-16': 'utf-16'
        }
        valid_encoding = encoding_map.get(detected_encoding.lower(), 'utf-8')
        encoding_cell.value = valid_encoding
        encoding_cell.border = border_thin

        # Add encoding validation with enhanced settings
        encoding_validation = DataValidation(type="list", formula1='"utf-8,utf-16,iso-8859-1,cp1252"', showDropDown=False)
        encoding_validation.error = "Please select a valid encoding: utf-8, utf-16, iso-8859-1, or cp1252"
        encoding_validation.errorTitle = "Invalid Encoding"
        encoding_validation.prompt = "Click dropdown arrow to select file encoding"
        encoding_validation.promptTitle = "Encoding Selection"
        encoding_validation.showErrorMessage = True
        encoding_validation.showInputMessage = True
        ws.add_data_validation(encoding_validation)
        encoding_validation.add(encoding_cell)
        row += 1

        # Text Qualifier field
        ws[f'A{row}'].value = "Text Qualifier:"
        ws[f'A{row}'].font = label_font
        qualifier_cell = ws[f'B{row}']
        qualifier_cell.value = format_data.get('text_qualifier', '"')
        qualifier_cell.border = border_thin

        # Add text qualifier validation with enhanced settings
        qualifier_validation = DataValidation(type="list", formula1='"\",\',None"', showDropDown=False)
        qualifier_validation.error = "Please select a valid text qualifier: double quote (\"), single quote ('), or None"
        qualifier_validation.errorTitle = "Invalid Text Qualifier"
        qualifier_validation.prompt = "Click dropdown arrow to select text qualifier"
        qualifier_validation.promptTitle = "Text Qualifier Selection"
        qualifier_validation.showErrorMessage = True
        qualifier_validation.showInputMessage = True
        ws.add_data_validation(qualifier_validation)
        qualifier_validation.add(qualifier_cell)
        row += 1

        # Has Headers dropdown
        ws[f'A{row}'].value = "Has Headers:"
        ws[f'A{row}'].font = label_font
        headers_cell = ws[f'B{row}']
        headers_cell.value = "Yes" if format_data.get('has_header', True) else "No"
        headers_cell.border = border_thin

        # Add headers validation with enhanced settings
        headers_validation = DataValidation(type="list", formula1='"Yes,No"', showDropDown=False)
        headers_validation.error = "Please select Yes or No for headers"
        headers_validation.errorTitle = "Invalid Headers Option"
        headers_validation.prompt = "Click dropdown arrow: Yes if first row contains column names, No otherwise"
        headers_validation.promptTitle = "Headers Selection"
        headers_validation.showErrorMessage = True
        headers_validation.showInputMessage = True
        ws.add_data_validation(headers_validation)
        headers_validation.add(headers_cell)

        ws[f'C{row}'].value = "First row contains column names"
        row += 1

        return row

    def _add_processing_section(self, ws, csv_path: str, row: int, section_fill, section_font, label_font, border_thin) -> int:
        """Add processing configuration section with dropdowns"""

        # Section header
        ws.merge_cells(f'A{row}:C{row}')
        header_cell = ws[f'A{row}']
        header_cell.value = "Processing Configuration"
        header_cell.fill = section_fill
        header_cell.font = section_font
        header_cell.alignment = Alignment(horizontal='center')
        row += 1

        # Processing Mode
        ws[f'A{row}'].value = "Processing Mode:"
        ws[f'A{row}'].font = Font(bold=True, size=11, color="2F5597")
        row += 1

        ws[f'A{row}'].value = "Mode:"
        ws[f'A{row}'].font = label_font
        mode_cell = ws[f'B{row}']
        mode_cell.value = "Select Mode"  # Prompt user to select
        mode_cell.border = border_thin
        mode_cell.fill = PatternFill(start_color="FFEB3B", end_color="FFEB3B", fill_type="solid")  # Yellow highlight

        # Add processing mode validation with enhanced settings
        mode_validation = DataValidation(type="list", formula1='"fullload,append"', showDropDown=False)
        mode_validation.error = "Please select processing mode: fullload (truncates existing data) or append (adds to existing data)"
        mode_validation.errorTitle = "Processing Mode Required"
        mode_validation.prompt = "REQUIRED: Click dropdown arrow to select fullload or append mode"
        mode_validation.promptTitle = "Processing Mode Selection"
        mode_validation.showErrorMessage = True
        mode_validation.showInputMessage = True
        mode_validation.allowBlank = False
        ws.add_data_validation(mode_validation)
        mode_validation.add(mode_cell)

        ws[f'C{row}'].value = "fullload=truncate, append=add"
        row += 1

        # Add explanation row
        ws[f'A{row}'].value = ""
        ws[f'B{row}'].value = "• fullload: Deletes existing data first"
        ws[f'B{row}'].font = Font(size=9, italic=True, color="666666")
        row += 1

        ws[f'A{row}'].value = ""
        ws[f'B{row}'].value = "• append: Adds to existing data"
        ws[f'B{row}'].font = Font(size=9, italic=True, color="666666")
        row += 1

        # Reference Data Table
        ws[f'A{row}'].value = "Reference Data Table:"
        ws[f'A{row}'].font = Font(bold=True, size=11, color="2F5597")
        row += 1

        ws[f'A{row}'].value = "Create Config Record:"
        ws[f'A{row}'].font = label_font
        ref_data_cell = ws[f'B{row}']
        ref_data_cell.value = "No"  # Default changed to No
        ref_data_cell.border = border_thin

        # Add reference data validation with enhanced settings
        ref_data_validation = DataValidation(type="list", formula1='"Yes,No"', showDropDown=False)
        ref_data_validation.error = "Please select Yes (create config record) or No (skip config record)"
        ref_data_validation.errorTitle = "Invalid Reference Data Option"
        ref_data_validation.prompt = "Click dropdown arrow: Yes to create config record, No to skip"
        ref_data_validation.promptTitle = "Reference Data Configuration"
        ref_data_validation.showErrorMessage = True
        ref_data_validation.showInputMessage = True
        ws.add_data_validation(ref_data_validation)
        ref_data_validation.add(ref_data_cell)

        ws[f'C{row}'].value = "Create entry in ref.Reference_Data_Cfg"
        row += 1

        # Table Name
        ws[f'A{row}'].value = "Table Name:"
        ws[f'A{row}'].font = label_font
        table_name_cell = ws[f'B{row}']
        
        # Derive default table name from CSV filename
        csv_file = Path(csv_path)
        default_table_name = csv_file.stem.replace('-', '_').replace(' ', '_')
        table_name_cell.value = default_table_name  # Default derived from filename
        table_name_cell.border = border_thin

        ws[f'C{row}'].value = f"Default: {default_table_name} (derived from filename)"
        row += 1

        return row

    def _add_confirmation_section(self, ws, row: int, warning_fill, warning_font, label_font, border_thin) -> int:
        """Add final confirmation section"""

        # Section header
        ws.merge_cells(f'A{row}:C{row}')
        header_cell = ws[f'A{row}']
        header_cell.value = "FINAL CONFIRMATION (Required)"
        header_cell.fill = warning_fill
        header_cell.font = warning_font
        header_cell.alignment = Alignment(horizontal='center')
        row += 1

        # Important notice
        ws.merge_cells(f'A{row}:C{row}')
        notice_cell = ws[f'A{row}']
        notice_cell.value = "IMPORTANT: Please review all settings above carefully before confirming."
        notice_cell.font = Font(bold=True, size=11, color="FF0000")
        notice_cell.alignment = Alignment(horizontal='center')
        row += 1

        # Confirmation dropdown
        ws[f'A{row}'].value = "I Confirm Processing:"
        ws[f'A{row}'].font = Font(bold=True, size=11, color="FF0000")
        confirm_cell = ws[f'B{row}']
        confirm_cell.value = "No"  # Default - must be changed by user
        confirm_cell.border = border_thin
        confirm_cell.fill = PatternFill(start_color="FFEB3B", end_color="FFEB3B", fill_type="solid")

        # Add confirmation validation with enhanced settings
        confirm_validation = DataValidation(type="list", formula1='"No,Yes"', showDropDown=False)
        confirm_validation.error = "REQUIRED: You must select 'Yes' to authorize file processing"
        confirm_validation.errorTitle = "Final Confirmation Required"
        confirm_validation.prompt = "CRITICAL: Click dropdown arrow and select 'Yes' to authorize processing"
        confirm_validation.promptTitle = "Final Processing Authorization"
        confirm_validation.showErrorMessage = True
        confirm_validation.showInputMessage = True
        ws.add_data_validation(confirm_validation)
        confirm_validation.add(confirm_cell)

        ws[f'C{row}'].value = "Must be 'Yes' to process"
        ws[f'C{row}'].font = Font(color="FF0000", bold=True)
        row += 1

        # Processed by field
        ws[f'A{row}'].value = "Processed By:"
        ws[f'A{row}'].font = label_font
        processed_by_cell = ws[f'B{row}']
        processed_by_cell.value = ""  # User must fill
        processed_by_cell.border = border_thin
        processed_by_cell.fill = PatternFill(start_color="FFEB3B", end_color="FFEB3B", fill_type="solid")

        ws[f'C{row}'].value = "Required for audit trail"
        row += 1

        # Date/Time field (auto-filled but editable)
        ws[f'A{row}'].value = "Date/Time:"
        ws[f'A{row}'].font = label_font
        datetime_cell = ws[f'B{row}']
        datetime_cell.value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        datetime_cell.border = border_thin

        ws[f'C{row}'].value = "Processing timestamp"
        row += 2

        # Final warning
        ws.merge_cells(f'A{row}:C{row}')
        warning_cell = ws[f'A{row}']
        warning_cell.value = "⚠️ The file will NOT be processed until confirmation is 'Yes' and the Excel file is saved!"
        warning_cell.font = Font(bold=True, size=12, color="FF0000")
        warning_cell.alignment = Alignment(horizontal='center')
        warning_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        row += 1

        return row

    def _add_instructions_section(self, ws, row: int):
        """Add instructions section"""

        # Add some spacing
        row += 1

        # Instructions header
        ws.merge_cells(f'A{row}:C{row}')
        header_cell = ws[f'A{row}']
        header_cell.value = "📋 Instructions"
        header_cell.font = Font(bold=True, size=12, color="2F5597")
        header_cell.alignment = Alignment(horizontal='center')
        row += 1

        instructions = [
            "1. Review File Information to verify correct file details",
            "2. Check and modify CSV Format settings as needed",
            "3. Select appropriate Processing Mode and Reference Data options",
            "4. Fill in Table Name if different from filename",
            "5. Change 'I Confirm Processing' to 'Yes'",
            "6. Fill in your name in 'Processed By' field",
            "7. Save this Excel file to trigger processing",
            "",
            "💡 Tips:",
            "• Use dropdowns to select valid options",
            "• Yellow highlighted fields are required",
            "• Processing starts automatically after saving with confirmation='Yes'",
            "• Check logs for processing status and results"
        ]

        for instruction in instructions:
            ws[f'A{row}'].value = instruction
            if instruction.startswith(('💡', '•')):
                ws[f'A{row}'].font = Font(italic=True, size=10, color="666666")
            else:
                ws[f'A{row}'].font = Font(size=10)
            row += 1

    def populate_form_fields(self, detection_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Map CSV detection data to Excel form fields

        Args:
            detection_data: Output from CSVFormatDetector

        Returns:
            Dictionary of form field values
        """
        return {
            'delimiter': detection_data.get('column_delimiter', ','),
            'text_qualifier': detection_data.get('text_qualifier', '"'),
            'encoding': detection_data.get('encoding', 'utf-8'),
            'has_headers': detection_data.get('has_header', True),
            'processing_mode': 'Select Mode', # User must select
            'is_reference_data': 'No',       # Default changed to No
            'table_name': '',                # User to fill
            'confirmed': 'No'                # User must confirm
        }